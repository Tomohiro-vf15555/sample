import openpyxl
from pathlib import Path
import datetime

import tkinter as tk
from tkinter import filedialog

# Tkinterウィンドウの作成（非表示）
root = tk.Tk()
root.withdraw()

# ファイル選択ダイアログを表示し、選択されたファイルのパスを取得
file_path = filedialog.askopenfilename()

# ファイルが選択されなかった場合の処理
if not file_path:
    print("ファイルが選択されませんでした。")
else:
    print("選択されたファイル:", file_path)
f_name_s = Path(file_path).stem
indxs = f_name_s.find("組立図")
print("組立図:",f_name_s)
fn = f_name_s[:indxs]#スライスで先頭4文字


#fn = input("保存するファイル名を入力")
fnx = f_name_s + ".xlsx"
fns = fn + "/"

#Exploreで選択したファイルを開く
wb_zrv = openpyxl.load_workbook(file_path, data_only = True)

file_zrv = fnx
#file_zrv = "TNY組立図.xlsx"
file_output = fn + "部番一覧"

#組立図ファイルを読み込む
#wb_zrv = openpyxl.load_workbook(file_zrv)
print("sheetsname",wb_zrv.sheetnames)
s_names = wb_zrv.sheetnames
s_name = s_names[0]
ws_zrv = wb_zrv[s_name]
#print(s_name)

#A列を検索して下段部分があるか識別する
#
cnt = 1
check = 0
delta = ""
for row in ws_zrv.iter_rows(min_row = 40):

    if row[0].value == "変更":
        print(f"hennkou :ok{cnt + 39}")
        check = cnt + 39
        delta = "Δ"
        break
    elif cnt > 50:
        print("no words")
        break
    cnt += 1

#下段があった場合は下段のデータも取得して表示する
bbn_list = []
bbn_list2 = []

for c_row in range(1,3):

    #部番、部品名を取り出してlistに入れる
    #checkの値をmin_rowに設定する
    r = 14
    if c_row == 2 and check >= 40:
        r = check
    elif c_row == 2 and check < 40:
        break#下段が無かったら取り出し処理を終了する

    for row in ws_zrv.iter_rows(min_row = r):
        #print("row[2]:",row[2].value)
        if row[2].value is None:
            break
    
        value_list = []
        bbn_s = ""
        for c in range(1,7):
            bn = row[c].value
        #print(row[2].value[7:13])
            if bn is not None:
                bbn = bn.replace(" ","")
                hi = bbn[9:10]
            else:
                bbn = ""
            if c == 2:
                #print(bbn[6:12],hi)
                bbn_s = bbn[6:12]
            value_list.append(bbn)
        print(bbn_s)
        value_list.insert(7,bbn_s)#G列の位置に中部番までを入れる
        #上下で保存するlistを変更する
        if r == 14: 
            bbn_list.append(value_list)
        else:
            bbn_list2.append(value_list)
            #print(bbn_list2)
    print(bbn_list)
    #print(start_row)



#部番用シートを追加
ws_new = wb_zrv.create_sheet(title = "部番一覧")
coment = "※文字列1と文字列2が結合されたデータになります"
ws_new.append(["No","部番","部品名",
               "SEC","品目","OP","文字列1","文字列2",coment])
#この次の行からbbn_listws_newに追加する

for idx, row in enumerate(bbn_list, start = 2):
    #ws_new.append([idx - 1] + row)
    ws_new.append(row)
    print(idx)

#部番用シートを追加
ws_new2 = wb_zrv.create_sheet(title = "部番一覧2")
coment = "※文字列1と文字列2が結合されたデータになります"
ws_new2.append(["No","部番","部品名",
               "SEC","品目","OP","文字列1","文字列2",coment])
#この次の行からbbn_listws_newに追加する


for idx, row in enumerate(bbn_list2, start = 2):
    #ws_new.append([idx - start_row + 1] + row)
    ws_new2.append(row)
#列幅の調整
#sheetを選択する
ws_s = [ws_new,ws_new2]
for s_number in ws_s:
    print(s_number.title) 
#for col in ws_new.columns:
    for col in s_number.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)

            except:
                pass

        adjusted_width = (max_length + 2) * 1.2
        #ws_new.column_dimensions[column].width = adjusted_width
        s_number.column_dimensions[column].width = adjusted_width
# G, H列の幅を10に変更するコードを追加する
    s_number.column_dimensions['G'].width = 10
    s_number.column_dimensions['H'].width = 10
nday=datetime.datetime.today()
new=nday.strftime("%m.%d")
print(new)
if delta == "Δ":
    wb_zrv.save(f"{file_output}Δ ({new}).xlsx")    
else:
    wb_zrv.save(f"{file_output}({new}).xlsx")
#wb_new.save(f"{file_output}({s_name}).xlsx")
print(f"{file_output}({new}).xlsx")
