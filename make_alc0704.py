import openpyxl
from pathlib import Path
import datetime

import tkinter as tk
from tkinter import filedialog

# Tkinterウィンドウの作成（非表示）
root = tk.Tk()
root.withdraw()

# ファイル選択ダイアログを表示し、選択されたファイルのパスを取得
file_path = filedialog.askopenfilename()

# ファイルが選択されなかった場合の処理
if not file_path:
    print("ファイルが選択されませんでした。")
else:
    print("選択されたファイル:", file_path)

#エクスプローラからファイルを選択(部番一覧)
#部番一覧のpathを取得する → wb_path
#部番一覧のsheet1,2を設定する

#wb_path = "ZR-V部番一覧(MP-016(MP)).xlsx"
wb_path = file_path
wb = openpyxl.load_workbook(wb_path)
s_names = wb.sheetnames
s_name = s_names[1]
s_name2 = s_names[2]
ws_itran = wb[s_name]

print(s_name)

#pathから機種名を取り出す
f_name_s = Path(file_path).stem
indxs = f_name_s.find("部番一覧")
print("部番一覧:",f_name_s,indxs)
fn = f_name_s[:indxs]#スライスで先頭4文字
chck = f_name_s[indxs +4]#ファイル名からΔを取り出す
print("chck=",chck)

#各データ収納用list
bname_list = []#部番一覧
bname_list2 = []
bname_op_list =[]
bbn_list = []#G,H列のALC情報
data_list = []#全bbn_list
data_list2 = []
#部番、ALC情報取り出し
#sheet2,sheet3からそれぞれ部番、ALC情報を別のlistに取り出す
for k in range(1,3):
    s_namex = s_names[k]
    ws_itran = wb[s_namex]
    print("部番一覧sheet:",k,s_namex)
    for bn in ws_itran.iter_rows(min_row = 3):#部番取得開始を3行目に変更
        b_name = bn[1].value#部番を取り出し
        op = bn[5].value
        if b_name is None:
            break
        bname_op_list = [b_name,op]
        if k == 1:
            bname_list.append(bname_op_list)
        else:
            bname_list2.append(bname_op_list)
        #bname_list.append(b_name)#b_name(部番)を入れる
        
        bbn_list=[]#初期化
        i = 6
        #G,H列のALCの文字列をループで取得する
        for i in range(6,8):
            #入力有無判定、入力が無い場合は''を設定する
            if bn[i].value is not None:
                alc1 = bn[i].value
            else:
                alc1 = ''
            bbn_list.append(alc1)#部番別ALC表示内容listにALC文字列を入れる
        if k == 1:
            data_list.append(bbn_list)#全bbn_listへbbn_listを入れる
        else:
            data_list2.append(bbn_list)
print("全ALC情報:data_list2")
print(data_list2)
#print("部番一覧:bname_list")
#print(bname_list)
print("import model_list_explorer0701")

#ここでmodel_list.pyを呼び出して実行させる。
import model_list_explorer0701

model_list_explorer0701.make_model_list(file_path)

#code_pathに"**機種データ.xlsx"を代入する
code_path = model_list_explorer0701.make_model_list(file_path)
print(code_path,"受け取り")
#実行してできた機種データ.xlsxを開く
wb_code = openpyxl.load_workbook(code_path)
code_names = wb_code.sheetnames
#wb_codeのSheet1を検索して上下段を識別する
s_names = wb_code.sheetnames
s_name = s_names[0]
ws_code0 = wb_code[s_name]
print(s_name)

#A列を検索して下段部分があるか識別する
#
cnt = 1
check = 0
for row in ws_code0.iter_rows(min_row = 40):

    if row[0].value == "変更":
        print(f"hennkou :ok{cnt + 39}")
        print(row[6].value,row[7].value)
        
        check = cnt + 39
        break
    elif cnt > 50:
        print("no words")
        break
    cnt += 1
#cnt_b = check

#chckによって処理をループしてデータ選択を切り替える
chenge_d = 0
for chenge_d in range(2):
    #1回目と2回目でsheetを変更して処理を分ける
    if chenge_d == 0:
        #sheet3を選択する
        c_name = code_names[3]
        bname_listx = bname_list
        data_listx = data_list
    elif chenge_d != 0 and check > 20:
        c_name = code_names[4]
        bname_listx = bname_list2
        data_listx = data_list2
    #print("chck=",check,chenge_d,c_name)
    ws_code = wb_code[c_name] #sheet3のセルのデータ
    #print(f"sheet3:{c_name[-5:]}")
    #print("data_list=",data_listx)

    #機種データsheet 6. 14行目のH列以降のデータを取得 (派生入力行のデータall)
    row_14_data = list(ws_code.iter_rows(min_row=4, max_row=4, min_col=9, values_only=False))[0]
    #print("派生行データ一覧",row_14_data)

    #派生入力行の データが入力されている列のインデックスを取得
    columns_with_data = []#入力ありの列indexのリスト
    for col_idx, cell in enumerate(row_14_data, start=1):#start=8→start=1
        if cell.value is not None:
            columns_with_data.append(col_idx)#列のindexを入れる
    #print("派生入力セルの列index",columns_with_data)
    #機種コード行別に取得
    k_data_list = []#全データlist
    hasei_list = []#派生別list
    tekiyo_list = []#部品別適用データlist

    for k_data in ws_code.iter_rows(min_row = 9):
        hasei_list = []
        kisyu = k_data[1].value#機種コード
        hasei = k_data[7].value#派生コード
        print(hasei)
        if kisyu is None:
            break
        
        #適用欄を列方向に処理
        tekiyo_list = []#結合したALC情報を入れるlist
        tekiyo_listop = []#結合したALC情報を入れるlist(op)
        for index in columns_with_data:
            idx = index + 7#H列からになるように変数に+7
            print(idx,index)
            tekiyo = k_data[idx]#適用欄の値
            print(hasei,"適用欄:",tekiyo.value)
            #適用欄の値がある場合に処理
            if tekiyo is not None and index - 1 < len(row_14_data):
                #indexのセルを指定する
                cell = ws_code.cell(row = k_data[0].row,column = index + 8).value#適用欄セル
                b_cell = ws_code.cell(4, index + 8).value#部番セル
                op_cell = ws_code.cell(8, index + 8).value#OP記号セル
                print(f"cell=,{b_cell},適用欄:{cell},OP:{op_cell},行:{k_data[0].row},列:{index + 8}")
                if 0 < index - 1 <= len(bname_listx):
                    
                    
                    #部番を照合する(同じindexでデータをとる)
                    #bbn_list側の部番、OP記号を取り出す
                    bbn = bname_listx[index -1]
                    bbn_n = bbn[0]#部番
                    bbn_o = bbn[1]#OP記号
                    print("部番照合:",bbn_n,b_cell)
                else:
                    print(f"Error: index {index - 1} out of range for bname_list with length {len(bname_listx)}")
                    bbn_n = None
                    bbn_o = None
                if bbn_n != b_cell:
                    print("部番照合NG")
                else:
                    print("部番照合◎")
                s_cell = str(cell)
                #op入力識別用変数s_cell
                if len(str(cell)) == 2:
                    s_cell = s_cell[:1]
                print("OP1文字:",s_cell)

                print("OP照合",op_cell,bbn_o)
                kisyu_op = ""
                hasei_op = ""
                #部番一覧のOP記号が"_"以外の場合に照合する
                if bbn_o != "_" and bbn_o == op_cell:
                    kisyu_op = kisyu#機種コードop
                    hasei_op = hasei#派生コードop
                    print("OP照合◎",bbn_o,str(cell),kisyu_op)
                elif bbn_o != op_cell:
                    print("OP照合NG")
                print("ALC",data_listx[index - 1])
                
                #cellに入力値があるか判定する
                alc_3 = ""
                alc_4 = ""
                if cell is not None and index - 1 < len(data_list):
                    alc_a = data_listx[index - 1]
                    print("ALC情報:",alc_a)
                    #ALC情報の文字列を取り出してtekiyo_listに入れる
                    for i in range(2):
                        #ループ1回目はG列のデータ(i = 0)
                        if  i == 0 and alc_a[i] is not None:
                            #通常ALC　適用欄 is not ,op_cell="_"の時だけ取得
                            if s_cell != "F":
                                alc_1 = alc_a[0]
                            #kisyu_OP(OP記号)の有,適用欄=f1,op_cell!="_"で変数を切り替える
                            if kisyu_op is not None and s_cell == "F":
                                alc_3 = alc_a[0]
                
                            
                            #print("?",alc_a[0],alc_1)
                        #ループ2回目はH列(i = 1)
                        elif i == 1 and alc_a[i] is not None:
                            if s_cell != "F":
                                alc_2 = alc_a[1]
                            if kisyu_op is not None and s_cell == "F":
                                alc_4 = alc_a[1]
                            
                            
                            #print("!",alc_a[1],alc_2)
                    #通常のALC情報をtekiyo_listに入れる
                    if alc_1 or alc_2:
                        alc_plus = "".join([alc_1,alc_2])
                        
                        tekiyo_list.append(alc_1)
                        tekiyo_list.insert(1,alc_2)
                        #tekiyo_list.append(alc_plus)
                        alc_plus = ""

                    #print("ALC1,2:",alc_a,alc_1,alc_2)
                    #print(alc_plus)
                    print("tekiyo_list:",tekiyo_list)
                    print("alc_3 =",alc_3,len(alc_3))
                    #ＯＰ記号のALC情報をtekiyo_listopに入れる
                    if alc_3 is not None or alc_4 is not None:
                        
                        tekiyo_listop.append(alc_3)
                        tekiyo_listop.insert(1,alc_4)
                    else:
                        continue
                    #tekiyo_listopから op_value_listに入れ直す
                    op_value_list = []
                    cnt = -1
                    #逆順のデータを後ろから取り出して未入力を削除する
                    for op_value in tekiyo_listop:
                        if op_value != "":
                
                            op_value_list.insert(cnt,op_value)
                            cnt -= 1

                    print("op_value =",op_value_list)
                    #print("tekiyo_list_op:",tekiyo_listop)
        #派生毎のデータをk_data_listにまとめる
        hasei_list.append(tekiyo_list)
        k_data_list.append([kisyu,hasei,""] + tekiyo_list)
        
        #bbn_o != "_" and cell is not NOne の時op記号のデータを追加
        if bbn_o != "_" and cell is not None:
            k_data_list.append([kisyu_op,hasei_op,bbn_o] + op_value_list)
         
        tekiyo_list = []  
        
    #データをコピー
    
    ws_new = wb_code.create_sheet(title = "ALC_data")

    ws_new.append(["機種","派生","","文字列1","文字列2","ALC"])
    for alc_data in k_data_list:
        ws_new.append(alc_data)
        s = len(k_data_list)

    for row  in range(2,s + 2):
        c_value = ws_new.cell(row,1).value
        if c_value is None:
            break
        else:
            a1_value = ws_new.cell(row,4).value
            a2_value = ws_new.cell(row,5).value
            a1 = str(a1_value)if a1_value is not None else ""
            a2 = str(a2_value)if a2_value is not None else ""
            a_j = a1 + a2
            ws_new.cell(row,6).value = a_j
            #print(c_value)
    #print(s)
#日付設定とファイルの保存
nday=datetime.datetime.today()
new=nday.strftime("%m.%d")
wb_code.save(f"ALC({fn}){new}.xlsx")
print(f"ALC({fn}){new}.xlsx")