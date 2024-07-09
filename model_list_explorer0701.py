import openpyxl
from pathlib import Path
import datetime

import tkinter as tk
from tkinter import filedialog

#関数化
def make_model_list(file_path):
    print("model_list.py")
    
    f_name_s = Path(file_path).stem
    indxs = f_name_s.find("部番一覧")
    print("部番一覧:",f_name_s,indxs)
    fn = f_name_s[:indxs]#スライスで先頭4文字

    #部番一覧ファイルを開く
    wb_zrv = openpyxl.load_workbook(file_path, data_only = True)

    #保存するファイル名を設定
    file_output = fn + "機種データ0701.xlsx"
    #sheet1のsheet名を取得して選択する
    s_names = wb_zrv.sheetnames
    s_name = s_names[0]
    ws_zrv = wb_zrv[s_name]

    #A列を検索して下段部分があるか識別する
    #
    cnt = 1
    check = 0
    for row in ws_zrv.iter_rows(min_row = 40):

        if row[0].value == "変更":
            print(f"hennkou :ok{cnt + 39}")
            print(row[6].value,row[7].value)
            
            check = cnt + 39
            break
        elif cnt > 50:
            print("no words")
            break
        cnt += 1
    cnt_b = check
    #print("sikibetu//",ws_zrv.cell(check - 6,8).value)
    #checkが0以上の場合、セル結合解除の下限を設定する
    if cnt_b > 0:#check > 0:
        #lower_limit = max(check - 10,1)#下限最小値を1
        lower_limit = cnt_b#check#下限を下段の項目行に設定
    elif cnt_b == 0:
        lower_limit = 60
    print("lowerlmit=",lower_limit)


    # 1. 結合されたセルの範囲を取得し、一時的にリストに保存する
    merged_cell_ranges = list(ws_zrv.merged_cells.ranges)

    # 2. 結合セルの内容を保存
    merged_cell_values = {}
    for merged_cell_range in merged_cell_ranges:
        # 結合セル範囲の左上のセルを取得
        top_left_cell = ws_zrv[merged_cell_range.coord.split(':')[0]]
        top_left_value = top_left_cell.value[:3] if top_left_cell.value else None
        # 結合範囲内のすべてのセルをループし、値を保存
        for row in ws_zrv[merged_cell_range.min_row:merged_cell_range.max_row + 1]:
            for cell in row[merged_cell_range.min_col - 1:merged_cell_range.max_col]:
                merged_cell_values[cell.coordinate] = top_left_value



    # データ確認
    print("Before unmerge:")
    print(ws_zrv['H8'].value)
    print(ws_zrv['I8'].value)
    print(ws_zrv['CP8'].value)
    print(ws_zrv['H47'].value)
    print(ws_zrv['I47'].value)
    print(ws_zrv['H53'].value)
    print(ws_zrv['I53'].value)

    # 3. 取得した範囲の結合を解除する
    for merged_cell_range in merged_cell_ranges:
        #print("gyou:",merged_cell_range)
        #print(merged_cell_range.min_row,merged_cell_range.min_col,merged_cell_range.max_col)
        m_row = merged_cell_range.min_row
        m_col = merged_cell_range.min_col
        min_r,max_r = str(merged_cell_range).split(":")
        #print("str:",min_r,max_r)
        if merged_cell_range.max_row <= lower_limit:
            top_left_cell = ws_zrv[merged_cell_range.coord.split(':')[0]]
            top_left_value = top_left_cell.value[:3] if top_left_cell.value else None
            ws_zrv.unmerge_cells(str(merged_cell_range))
            
            # 結合範囲内のすべてのセルに左上の値を設定
            for row in range(merged_cell_range.min_row, merged_cell_range.max_row + 1):
                for col in range(merged_cell_range.min_col, merged_cell_range.max_col + 1):
                    ws_zrv.cell(row, col).value = top_left_value
                    #print(f"セル ({row}, {col}) に値を設定: {top_left_value}")

    #for cell_coord, value in merged_cell_values.items():
    #    ws_zrv[cell_coord] = value

    # データ確認
    print("\nAfter unmerge:")
    print(ws_zrv['H8'].value)
    print(ws_zrv['I8'].value)
    print(ws_zrv['CP8'].value)
    print(ws_zrv['H47'].value)
    print(ws_zrv['I47'].value)
    print(ws_zrv['H53'].value)
    print(ws_zrv['I53'].value)


    # 6. 14行目のH列以降のデータを取得 (min_col=8→min_col=1)

    #下段を識別したら行の設定を変更する
    #ループを設定して下段のデータも取得
    #行設定row = 14 上段= 14,下段= check 
    #行設定row = 7  上段= 7,下段= check- 7
    #行設定colmun_with_data(row_14_dataから取り出し)
    for i in range(2):
        print(f"ループ {i + 1} 回目の開始")
        if i == 1 and check < 14:
            print("check<14によるループ中断")
            break
        
        print("i = ",i)
        #check = 0 変数は基本設定
        if i == 0:
            s_row = 14
            ksu_row = 7
            if check > 14:
                max_row = check - 9
            else:
                max_row = ws_zrv.max_row
            k_code = ws_zrv.cell(8,8).value
        #check>0 ループ2回目は変数を変更する
        else:
            s_row = check
            ksu_row = check - 7
            max_row = ws_zrv.max_row
            k_code = ws_zrv.cell(check-6,8).value
            print("下段:",cnt_b,check)
        #print(f"機種:{k_code}")
        #print(f"変数切替: s_row = {s_row}, ksu_row = {ksu_row}")
        row_14_data = list(ws_zrv.iter_rows(min_row=s_row, max_row=s_row, min_col=1, values_only=False))[0]
        #print(f"row_14_data = {row_14_data}")

        # データが入力されている列のインデックスを取得
        columns_with_data = []
        #columns_with_data2 = []

        for col_idx, cell in enumerate(row_14_data, start=1):#行設定row = 14
            if cell.value is not None:
                columns_with_data.append(col_idx)
        print("columns_with_data?")
        #print(f"columns_with_data = {columns_with_data}")

        n = str(i + 1)
        ws_new = wb_zrv.create_sheet(title = k_code + "機種データ")

        # 項目名を追加
        ws_new.append(["No", "機種", "", "", "", "", "", "派生", "部番･適用"])

        # C列のインデックスを特定
        index_C_column = columns_with_data[2] - 1  # C列に対応するインデックス（0ベース）
        # 7行目以降のデータを取得
        zrv_data = []

        for row in ws_zrv.iter_rows(min_row=ksu_row, max_row = max_row, values_only=False):#行設定row = 7
        
            filtered_row = []
            for col_idx in columns_with_data:
                cell = ws_zrv.cell(row=row[0].row, column=col_idx)#行設定colmun_with_data
                value = cell.value
                
                if col_idx - 1 == index_C_column and value:
                    value = value.replace(" ","")

                filtered_row.append(value)
            

            zrv_data.append(filtered_row)
            #print(f"row: {row[0].row}, filtered_row: {filtered_row}")
            #print(f"row: {row[0].row}, column:{col_idx}value: {value}")
        # 行列を入れ替える
        transposed_data = list(zip(*zrv_data))

        # データのコピー
        #sheetを追加する
        #部番用シートを追加
        
        
        for row_idx, row in enumerate(transposed_data, start=2):
            for col_idx, value in enumerate(row, start=1):
                ws_new.cell(row=row_idx, column=col_idx, value=value)

    print(f"ループ {i + 1} 回目の終了")   
    # 新しいファイルに保存
    wb_zrv.save(file_output)
    
    print(f"\n{file_output} saved.")
    return file_output  # 処理結果のファイルパスを返す
   

    